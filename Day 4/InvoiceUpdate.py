from datetime import date
from openpyxl import Workbook, load_workbook
import os

# ================= UPDATE CUSTOMER DETAILS =================
def update_customer_details(cusDetails):
    # 1. Update option
    while True:
        print("""
        Update Customer Details:
        1 - Update Name
        2 - Update Phone Number
        3 - Update Email
        4 - Show Current Details
        0 - Exit Update Menu
        """)

        ch = input("Enter choice: ")

        if ch == '1':
            cusDetails[0] = input("Enter new name: ")
            print("Name updated!")

        elif ch == '2':
            phone = input("Enter new phone number: ")
            while (not phone.isdigit()) or len(phone) != 10 or int(phone[0]) < 6:
                phone = input("Enter valid 10-digit phone number starting with 6,7,8,9: ")
            cusDetails[1] = phone
            print("Phone number updated!")

        elif ch == '3':
            cusDetails[2] = input("Enter new email: ")
            print("Email updated!")

        elif ch == '4':
            print("\nCurrent Customer Details:")
            print("Name:", cusDetails[0])
            print("Phone:", cusDetails[1])
            print("Email:", cusDetails[2])

        elif ch == '0':
            break

        else:
            print("Invalid option!")


# ================= SHOW INVOICE =================
def show(items, TotalPrice, productCtr, cusDetails):
    print("\n\n:::: Invoice Details ::::")
    print(f"{cusDetails[0]}\t\t{date.today()}")
    print(f"{cusDetails[1]}\t\t{cusDetails[2]}")
    print("---------------------------------")
    print("Id\tName\tQty\tMRP")
    
    for i in range(productCtr):
        print(f"{items[i][0]}\t{items[i][1]}\t{items[i][2]}\t{items[i][3]}")
    
    print("---------------------------------")
    print("Total Price:", TotalPrice)

# ================= ADD PRODUCT =================
def addProduct(items, TotalPrice, productCtr):
    productId = input("Enter product ID: ")
    productName = input("Enter product Name: ")

    qty = input("Enter qty: ")
    while not qty.isdigit():
        qty = input("Enter valid qty: ")
    qty = int(qty)

    price = input("Enter MRP: ")
    while not price.isdigit():
        price = input("Enter valid MRP: ")
    price = int(price)

    items.append([productId, productName, qty, price])
    TotalPrice += qty * price
    productCtr += 1

    return items, TotalPrice, productCtr

# ================= DELETE FROM CURRENT LIST =================
def delete(items, productCtr, TotalPrice, productId):
    for i in items:
        if i[0] == productId:
            TotalPrice -= i[2] * i[3]
            items.remove(i)
            productCtr -= 1
            print("Deleted from current list.")
            return items, productCtr, TotalPrice

    print("Product not found!")
    return items, productCtr, TotalPrice

# ================= SAVE TO EXCEL =================
def save_to_excel(filename, items, TotalPrice, cusDetails):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.append(["Invoice","",date.today()])
    ws.append(["Name", cusDetails[0]])
    ws.append(["Contact", cusDetails[1]])
    ws.append(["Email", cusDetails[2]])
    ws.append([])
    ws.append(["ProductID", "ProductName", "Qty", "MRP"])

    for item in items:
        ws.append(item)

    ws.append([])
    ws.append(["Total Amount", TotalPrice])

    wb.save(filename)
    print(f"Invoice saved to {filename}")

# ================= DELETE FROM EXCEL =================
def delete_item_from_excel(filename, product_id):
    if not os.path.exists(filename):
        print("Excel file does not exist!")
        return

    wb = load_workbook(filename)
    ws = wb["Invoice"]

    rowNumber = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == product_id:
            rowNumber = row
            break

    if rowNumber:
        ws.delete_rows(rowNumber)
        print(f"Product {product_id} deleted from Excel.")
    else:
        print("Product ID not found in Excel!")

    wb.save(filename)

# ================= CUSTOMER DETAILS =================
cusDetails = []
cusDetails.append(input("Enter customer name: "))

phone = input("Enter contact number: ")
while (not phone.isdigit()) or len(phone) != 10 or int(phone[0]) < 6:
    phone = input("Enter valid 10-digit phone number starting with 6-9: ")

cusDetails.append(phone)
cusDetails.append(input("Enter email: "))

# ================= MAIN DATA =================
items = []
TotalPrice = 0
productCtr = 0
filename = "Invoice2.xlsx"

# ================= MENU LOOP =================
while True:
    choice = input("""
0 - Add product
1 - Show invoice
2 - Delete product from current list
3 - Update Customer details
5 - Delete product from Excel
6 - Save invoice to Excel
q - Exit
Enter choice: """)

    if choice == '0':
        items, TotalPrice, productCtr = addProduct(items, TotalPrice, productCtr)

    elif choice == '1':
        show(items, TotalPrice, productCtr, cusDetails)

    elif choice == '2':
        pid = input("Enter product ID to delete: ")
        items, productCtr, TotalPrice = delete(items, productCtr, TotalPrice, pid)

    elif choice=='3':
        # Funtion to edit the customer details like Name,Phone number,email
        update_customer_details(cusDetails)

    elif choice == '5':
        pid = input("Enter product ID to delete from Excel: ")
        delete_item_from_excel(filename, pid)

    elif choice == '6':
        save_to_excel(filename, items, TotalPrice, cusDetails)

    elif choice.lower() == 'q':
        break

    else:
        print("Invalid choice!")