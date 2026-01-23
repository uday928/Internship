from datetime import date
from openpyxl import Workbook, load_workbook
import os

SHOP_STATE = input("Enter state buyer state name:")   
GST_RATE = 0.18


# ================= UPDATE CUSTOMER DETAILS =================
def update_customer_details(cusDetails):
    while True:
        print("""
        Update Customer Details:
        1 - Update Name
        2 - Update Phone Number
        3 - Update Email
        4 - Update State
        5 - Show Current Details
        0 - Exit Update Menu
        """)

        ch = input("Enter choice: ")

        if ch == '1':
            cusDetails[0] = input("Enter new name: ")
            print("Name updated!")

        elif ch == '2':
            phone = input("Enter new phone number: ")
            while (not phone.isdigit()) or len(phone) != 10 or int(phone[0]) < 6:
                phone = input("Enter valid 10-digit phone number starting with 6-9: ")
            cusDetails[1] = phone
            print("Phone updated!")

        elif ch == '3':
            cusDetails[2] = input("Enter new email: ")
            print("Email updated!")

        elif ch == '4':
            cusDetails[3] = input("Enter new state: ")
            print("State updated!")

        elif ch == '5':
            print("\nCustomer Details:")
            print("Name:", cusDetails[0])
            print("Phone:", cusDetails[1])
            print("Email:", cusDetails[2])
            print("State:", cusDetails[3])

        elif ch == '0':
            break
        else:
            print("Invalid choice!")


# ================= SHOW INVOICE =================
def show(items, TotalPrice, productCtr, cusDetails):
    print("\n\n:::: INVOICE DETAILS ::::\n")
    print(f"Customer: {cusDetails[0]}        Date: {date.today()}")
    print(f"Phone: {cusDetails[1]}    Email: {cusDetails[2]}")
    print(f"State: {cusDetails[3]}")
    print("-" * 80)
    print("ID  Name   Qty  MRP  CGST   SGST   IGST   TOTAL")

    for i in range(productCtr):
        p = items[i]
        print(f"{p[0]} {p[1]} {p[2]} {p[3]} {p[4]:.2f} {p[5]:.2f} {p[6]:.2f} {p[7]:.2f}")

    print("-" * 80)
    print("GRAND TOTAL =", TotalPrice)


# ================= ADD PRODUCT =================
def addProduct(items, TotalPrice, productCtr, cusDetails):
    productId = input("Enter product ID: ")
    productName = input("Enter product Name: ")

    qty = input("Enter qty: ")
    while not qty.isdigit():
        qty = input("Enter valid qty: ")
    qty = int(qty)

    price = input("Enter MRP: ")
    while not price.isdigit():
        price = input("Enter valid MRP: ")
    price = int(price)

    base_price = qty * price
    customer_state = cusDetails[3]

    # GST Calculation
    if customer_state.lower() == SHOP_STATE.lower():
        cgst = base_price * 0.09
        sgst = base_price * 0.09
        igst = 0
    else:
        cgst = 0
        sgst = 0
        igst = base_price * GST_RATE

    tax = cgst + sgst + igst
    total = base_price + tax

    items.append([productId, productName, qty, price, cgst, sgst, igst, total])
    TotalPrice += total
    productCtr += 1

    print("Product Added with GST!")
    return items, TotalPrice, productCtr


# ================= DELETE FROM CURRENT LIST =================
def delete(items, productCtr, TotalPrice, productId):
    for i in items:
        if i[0] == productId:
            TotalPrice -= i[7]
            items.remove(i)
            productCtr -= 1
            print("Deleted from current list.")
            return items, productCtr, TotalPrice

    print("Product not found!")
    return items, productCtr, TotalPrice


# ================= SAVE TO EXCEL =================
def save_to_excel(filename, items, TotalPrice, cusDetails):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.append(["INVOICE", "", date.today()])
    ws.append(["Name", cusDetails[0]])
    ws.append(["Contact", cusDetails[1]])
    ws.append(["Email", cusDetails[2]])
    ws.append(["State", cusDetails[3]])
    ws.append([])

    ws.append(["ProductID", "ProductName", "Qty", "MRP", "CGST", "SGST", "IGST", "Total"])

    for item in items:
        ws.append(item)

    ws.append([])
    ws.append(["Grand Total", TotalPrice])

    wb.save(filename)
    print("Invoice saved to", filename)


# ================= DELETE FROM EXCEL & UPDATE TOTAL =================
def delete_item_from_excel(filename, product_id):
    if not os.path.exists(filename):
        print("Excel file not found!")
        return

    wb = load_workbook(filename)
    ws = wb["Invoice"]

    rowNumber = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == product_id:
            rowNumber = row
            break

    if rowNumber:
        ws.delete_rows(rowNumber)
        print(f"Product {product_id} deleted from Excel.")

        # Recalculate total
        total = 0
        for row in range(8, ws.max_row):
            value = ws.cell(row=row, column=8).value
            if isinstance(value, (int, float)):
                total += value

        ws.cell(row=ws.max_row, column=2).value = total
        print("Excel Total Updated!")

    else:
        print("Product ID not found!")

    wb.save(filename)


# ================= CUSTOMER DETAILS INPUT =================
cusDetails = []
cusDetails.append(input("Enter customer name: "))

phone = input("Enter contact number: ")
while (not phone.isdigit()) or len(phone) != 10 or int(phone[0]) < 6:
    phone = input("Enter valid 10-digit phone starting with 6-9: ")
cusDetails.append(phone)

cusDetails.append(input("Enter email: "))
cusDetails.append(input("Enter State (e.g. Gujarat, Maharashtra): "))

# ================= MAIN DATA =================
items = []
TotalPrice = 0
productCtr = 0
filename = "Invoice_GST.xlsx"


# ================= MENU LOOP =================
while True:
    choice = input("""
0 - Add product
1 - Show invoice
2 - Delete product from current list
3 - Update customer details
5 - Delete product from Excel
6 - Save invoice to Excel
q - Exit
Enter choice: """)

    if choice == '0':
        items, TotalPrice, productCtr = addProduct(items, TotalPrice, productCtr, cusDetails)

    elif choice == '1':
        show(items, TotalPrice, productCtr, cusDetails)

    elif choice == '2':
        pid = input("Enter product ID to delete: ")
        items, productCtr, TotalPrice = delete(items, productCtr, TotalPrice, pid)

    elif choice == '3':
        update_customer_details(cusDetails)

    elif choice == '5':
        pid = input("Enter product ID to delete from Excel: ")
        delete_item_from_excel(filename, pid)

    elif choice == '6':
        save_to_excel(filename, items, TotalPrice, cusDetails)

    elif choice.lower() == 'q':
        break

    else:
        print("Invalid choice!")